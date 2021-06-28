#用oenpyxl处理Excel文件
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

wbo=load_workbook('资产估值表-太平资产吉祥2号货币型资管产品&工行-20210510.xlsx')
wso = wbo["太平资产吉祥2号货币型资管产品&工行"]
#print(wso.max_row)
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("统计结果", 0)
ws1['A1']='太平资产吉祥2号统计情况'
ws1['A2']='资产类别'
ws1['B2']='资产市值（元）'
ws1['C2']='资产占比(%)'
ws1['A3']='存款'
ws1['A4']='存单'
ws1['A5']='债券'
ws1['A6']='货币基金'
ws1['A7']='现金及回购'

for r in range(1,wso.max_row):
    if wso.cell(r,1).value == "1002.01":
         #print (r,wso.cell(r,2).value) #活期存款
         XJSZ=wso.cell(r,8).value
         XJBL=wso.cell(r,9).value
         ws1['A3']=wso.cell(r,2).value
         ws1['B3']=wso.cell(r,8).value
         ws1['C3']=wso.cell(r,9).value
         #ws1['B3'].number_format=u'#,##0.00'
         #ws1['C3'].number_format=u'#,##0.00'
         
    elif wso.cell(r,1).value == "1002.04": #协议存款
         #print (r,wso.cell(r,2).value)
         ws1['A4']=wso.cell(r,2).value
         ws1['B4']=wso.cell(r,8).value
         ws1['C4']=wso.cell(r,9).value
         #ws1['B4'].number_format=u'#,##0.00'
         #ws1['C4'].number_format=u'#,##0.00'   
    elif wso.cell(r,1).value == "1503.29": #同业存单
         #print (r,wso.cell(r,2).value)
         ws1['A5']=wso.cell(r,2).value
         ws1['B5']=wso.cell(r,8).value
         ws1['C5']=wso.cell(r,9).value
    elif wso.cell(r,1).value == "1103": #债券
         #print (r,wso.cell(r,2).value)
         ws1['A8']=wso.cell(r,2).value
         ws1['B8']=wso.cell(r,8).value
         ws1['C8']=wso.cell(r,9).value
    elif wso.cell(r,1).value == "1105": #货币基金
         #print (r,wso.cell(r,2).value)
         ws1['A6']=wso.cell(r,2).value
         ws1['B6']=wso.cell(r,8).value
         ws1['C6']=wso.cell(r,9).value
    elif wso.cell(r,1).value == "1202":  #回购
         #print (r,wso.cell(r,2).value)
         XJSZ=XJSZ+wso.cell(r,8).value
         XJBL=XJBL+wso.cell(r,9).value
         #print(XJSZ)
         #print(XJBL)
         ws1['A7']=wso.cell(r,2).value
         ws1['B7']=wso.cell(r,8).value
         ws1['C7']=wso.cell(r,9).value


#df=pd.read_excel('资产估值表-太平资产吉祥2号货币型资管产品&工行-20210510.xls',header=2,index_col=0)
#df1=df[['科目名称','市值','市值占净值比(%)']]
#print(df1)
#print(df1.index.values)
#for r in range(df1.index.values):
#    print (r,df1.index.values)

#XJ=df.loc["1002.01"].values
#CK=df.loc["1002.04"].values
#CD=df.loc["1503"].values
#ZQ=df.loc["1103"].values
#JJ=df.loc["1105"].values
#HG=df.loc["1202"].values

#ws1['B3']=CK[6]
#ws1['C3']=CK[7]
#ws1['B4']=CD[6]
#ws1['C4']=CD[7]
#ws1['B5']=ZQ[6]
#ws1['C5']=ZQ[7]
#ws1['B6']=JJ[6]
#ws1['C6']=JJ[7]
#ws1['B7']=XJ[6]+HG[6]
#ws1['C7']=XJ[7]+HG[7]

#ws1.number_format='Currency' 


ws1.column_dimensions['A'].width=18
ws1.column_dimensions['B'].width=18
ws1.column_dimensions['C'].width=15

for i in range(1,ws1.max_row+1):
    for j in range(1,ws1.max_column+1):
        ws1.cell(i,j).number_format=u'#,##0.00'

    

#ws1.column_dimensions['B'].number_format=u'#,##0.00'
#ws1.column_dimensions['C'].number_format=u'#,##0.00'

bold_14_font = Font(name='等线', size=18, color=colors.BLUE, bold=True)
bold_12_font = Font(name='等线', size=12, color=colors.BLUE, bold=True)
ws1['A1'].font = bold_14_font
ws1.merge_cells('A1:C1')




ws2 = wb.create_sheet("持仓情况",1)
ws2['A1']='太平资产吉祥2号持仓情况'
ws2['A2']='资产名词'
ws2['B2']='资产市值（元）'
ws2['C2']='资产占比(%)'
ws2.column_dimensions['A'].width=15
ws2.column_dimensions['B'].width=18
ws2.column_dimensions['C'].width=15
ws2['A1'].font = bold_14_font
ws2.merge_cells('A1:C1')
#ws2['A3']=df.loc["1002.01"].values[0]
#ws2['B3']=df.loc["1002.01"].values[6]
#ws2['C3']=df.loc["1002.01"].values[7]

#print(df.loc["1002.01"].市值)

#JX2 = load_workbook('资产估值表-太平资产吉祥2号货币型资管产品&工行-20210510.xlsx')
#print(JX2.sheetnames)

wb.save('太平资产吉祥2号统计结果.xlsx')

#import xlrd

#df=pd.read_excel('资产估值表-太平资产吉祥2号货币型资管产品&工行-20210510.xls',sheet_name=1,index_col=0)
#print(df)
#print(df.describe())

#df=pd.read_excel('资产估值表-太平资产吉祥2号货币型资管产品&工行-20210510.xls',header=2,index_col=0)
#print(df)
#print(df.loc["1002.01","1002.04""1103","1105","1202"],["市值","市值占净值比(%)"])
#print(df.loc["1002.01"])
#df.to_excel(excel_writer=r"测试输出.xls")



#for r in range(sheet.nrows):
#    print (r,sheet.row(r))
    
