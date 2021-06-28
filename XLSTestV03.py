import pandas as pd
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment



#df=pd.read_excel('资产估值表-太平资产吉祥2号货币型资管产品&工行-20210510.xls',header=2,index_col=0)
df=pd.read_excel('资产估值表-太平资产吉祥2号货币型资管产品&工行-20210510.xls',header=2)
#df1=df[['科目名称','市值','市值占净值比(%)']]
df1=df[['科目代码','科目名称','市值','市值占净值比(%)']]
#print(df1)
iHQ=df1.loc[df['科目代码']=='1002.01']
iXC=df1.loc[df['科目代码']=='1002.04']
iZQ=df1.loc[df['科目代码']=='1103']
iZPCB=df1.loc[df['科目代码']=='1103.73.01']
iZPSY=df1.loc[df['科目代码']=='1103.73.02']
iDRCB=df1.loc[df['科目代码']=='1103.74.01']
iDRSY=df1.loc[df['科目代码']=='1103.74.02']
iJJ=df1.loc[df['科目代码']=='1105']
iHJCB=df1.loc[df['科目代码']=='1105.01.01']
iHG=df1.loc[df['科目代码']=='1202']
iYHJHG=df1.loc[df['科目代码']=='1202.01.03']
iYSGL=df1.loc[df['科目代码']=='1203']
iCD=df1.loc[df['科目代码']=='1503.29']
iCDCB=df1.loc[df['科目代码']=='1503.29.01']
iCDSY=df1.loc[df['科目代码']=='1503.29.02']
iZCJZ=df1.loc[df['科目代码']=='净值(市值)']
iPLD=df1.loc[df['科目代码']=='偏离度']



#print(iZQ.index.values[0])

#df1=df1[(df1.index<iHQ.index.values[0])] #可筛选

#df1=df1[(df1.index>iXC.index.values[0])&(df1.index<iZQ.index.values[0])] 
#以下开始设置
dft=df1.iloc[iXC.index.values[0]+1:iZQ.index.values[0]]
df2=pd.concat([iHQ,dft])  #合并表格
dft=df1.iloc[iZPCB.index.values[0]+1:iZPSY.index.values[0]]
df2=pd.concat([df2,dft])
dft=df1.iloc[iDRCB.index.values[0]+1:iDRSY.index.values[0]]
df2=pd.concat([df2,dft])
dft=df1.iloc[iHJCB.index.values[0]+1:iHG.index.values[0]]
df2=pd.concat([df2,dft])
dft=df1.iloc[iYHJHG.index.values[0]+1:iYSGL.index.values[0]]
df2=pd.concat([df2,dft])
dft=df1.iloc[iCDCB.index.values[0]+1:iCDSY.index.values[0]]
df2=pd.concat([df2,dft])

#print(df2[0:10])
#df2=pd.concat([df2,iZCJZ]) #追加净值
#df2=pd.concat([df2,iPLD]) #追加偏离度

dfe=df2[['科目名称','市值','市值占净值比(%)']].reset_index(drop=True)
dft=dfe.sort_values(by=['市值'],ascending=[False])
df10=dft[0:10]

dfs=pd.concat([iHQ,iXC,iZQ,iCD,iJJ,iHG])
dfs=dfs[['科目名称','市值','市值占净值比(%)']].reset_index(drop=True)
#print(dfs)


writer=pd.ExcelWriter(".\太平资产吉祥2号统计结果.xlsx",engine="xlsxwriter")
dfs.to_excel(writer,sheet_name="持仓比例")
dft.to_excel(writer,sheet_name="持仓明细")
df10.to_excel(writer,sheet_name="前10大持仓")

writer.save()

#dfe.to_excel(excel_writer=r"太平资产吉祥2号统计结果02.xlsx",sheet_name="持仓明细")

#调整表格格式
wb = load_workbook('.\太平资产吉祥2号统计结果.xlsx')
ws1= wb["持仓比例"]
ws2= wb["持仓明细"]
ws3= wb["前10大持仓"]
ws1.delete_cols(1)
ws2.delete_cols(1)
ws3.delete_cols(1)
ws1.column_dimensions['A'].width=35
ws1.column_dimensions['B'].width=25
ws1.column_dimensions['C'].width=20
for i in range(1,ws1.max_row+1):
    for j in range(1,ws1.max_column+1):
        ws1.cell(i,j).number_format=u'#,##0.00'

bold_18_font = Font(name='等线', size=18, color=colors.BLUE, bold=True)
bold_12_font = Font(name='等线', size=12, color=colors.BLUE, bold=True)
ws1.insert_rows(1)
ws1['A1']="太平资产吉祥2号资产比例"
ws1['A1'].font = bold_18_font
ws1.merge_cells('A1:C1')


ws2.column_dimensions['A'].width=35
ws2.column_dimensions['B'].width=25
ws2.column_dimensions['C'].width=20
for i in range(1,ws2.max_row+1):
    for j in range(1,ws2.max_column+1):
        ws2.cell(i,j).number_format=u'#,##0.00'
ws2.insert_rows(1)
ws2['A1']="太平资产吉祥2号资产明细"
ws2['A1'].font = bold_18_font
ws2.merge_cells('A1:C1')

ws3.column_dimensions['A'].width=35
ws3.column_dimensions['B'].width=25
ws3.column_dimensions['C'].width=20
for i in range(1,ws2.max_row+1):
    for j in range(1,ws2.max_column+1):
        ws3.cell(i,j).number_format=u'#,##0.00'
ws3.insert_rows(1)
ws3['A1']="太平资产吉祥2号十大持仓"
ws3['A1'].font = bold_18_font
ws3.merge_cells('A1:C1')

wb.save('太平资产吉祥2号统计结果.xlsx')
wb.close()
print("文件:太平资产吉祥2号统计结果.xlsx，已经生成，请查收")
